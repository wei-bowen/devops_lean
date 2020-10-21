# _*_ coding=utf-8 _*_
# author:韦博文
# date:2019/10/23
# content:使用openpyxl操作excel文档

import openpyxl
import datetime


#获取一个workbook对象
'''
wb = openpyxl.load_workbook('test.xlsx')
print(wb.active)                            ##获取活跃的Worksheet
print(wb.read_only)                         ##判断是否只读模式打开
print(wb.encoding)                          ##获取文档的字符编码
print(wb.properties)                        ##文档的元数据
print(wb.sheetnames)                        ##列表返回所有sheet页

print('\n' +  '*' * 140 + '\n')

sheet1 = wb['Sheet1']                       ##根据sheet获取workersheet对象
'''
'''
sheet属性
'''
'''
print(sheet1.title)                         ##sheet标题
print(sheet1.dimensions)                    ##获取表格的数据范围
print(sheet1.max_row)                       ##最大行号
print(sheet1.min_row)                       ##最小行号
print(sheet1.max_column)                    ##最大列号
print(sheet1.min_column)                    ##最小列号
'''
'''
常用方法
iter_rows                                   :按行获取所有单元格(Cell对象)
iter_columns                                :按列获取所有单元格
append                                      :表格末尾添加数据
merged_cells                                :合并多个单元格
unmerge_cells                               :移除合并的单元格
'''
#print(list(sheet1.iter_rows(min_row=2,max_row=4,min_col=1,max_col=3)))
#print(sheet1['A1'].value)                   ##根据坐标获取单元格的值
#print(sheet1.cell(row=1,column=1).value)    ##按行列号获取值

'''
cell对象的属性：
row                                         ##单元格所在行
column                                      ##单元格所在列
value                                       ##单元格的值
cordinate                                   ##单元格坐标
'''

##遍历表格
#for row in sheet1.values :
#    print(*row)

#for row in sheet1.iter_rows() :
#    print(*[cell.value for cell in row])

'''
修改excel表格
'''
wb = openpyxl.Workbook()
ws = wb.active                                  #获取当前表格
ws.title = 'student'                            #修改表格名称
print(ws.title)

wb.create_sheet(index=0,title='new sheet')      #新建表格
print(wb.sheetnames)
wb.remove(wb['new sheet'])                      #删除表格
print(wb.sheetnames)
ws = wb.active
ws['A1'] = 'hello,world'                        #单元格赋值
ws['A2'] = datetime.datetime.now()
wb.save('sample.xlsx')                          #
